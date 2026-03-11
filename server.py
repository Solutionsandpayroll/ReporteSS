"""
Servidor local — Automatización SS
Ejecutar: python server.py
Abrir:    http://localhost:5000
"""
import io, os, copy, traceback, time
from datetime import date
from flask import Flask, request, send_file, jsonify, send_from_directory
import openpyxl
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
app  = Flask(__name__, static_folder=BASE)

# ─── Maestro ──────────────────────────────────────────────────────────────────
# URL de descarga directa del Excel en OneDrive. Configura la variable de entorno
# MAESTRO_URL con el enlace de compartir de OneDrive (1drv.ms o sharepoint.com).
# Si no se configura, se intenta leer maestro.xlsx local; si no existe, se usa
# el diccionario estático de abajo.
MAESTRO_URL = os.environ.get("MAESTRO_URL", "")

MAESTRO_CACHE_TTL = 300          # segundos entre recargas (5 min)
_maestro_cache      = None
_maestro_cache_time = 0.0


def _onedrive_to_download(url: str) -> str:
    """Convierte un enlace de compartir de OneDrive a URL de descarga directa."""
    import base64
    # OneDrive personal (1drv.ms / onedrive.live.com) → API de shares
    if "1drv.ms" in url or "onedrive.live.com" in url:
        token = base64.urlsafe_b64encode(("u!" + url).encode()).decode().rstrip("=")
        return f"https://api.onedrive.com/v1.0/shares/{token}/root/content"
    # OneDrive Business / SharePoint → añadir download=1
    if "sharepoint.com" in url:
        sep = "&" if "?" in url else "?"
        return url + sep + "download=1"
    return url  # asumir que ya es descarga directa


def _parse_maestro_xlsx(file_obj) -> dict:
    """Lee el Excel maestro de personal.

    Busca la fila de cabecera escaneando hasta la fila 10 (el archivo real
    tiene un título en fila 1 y la cabecera en fila 3). Las columnas se
    detectan por el texto del encabezado:
      CODIGO EMPLEADO → código  |  EPS → eps  |  AFP → afp
    Si no encuentra cabecera, usa posiciones conocidas: B=1, F=5, G=6.
    """
    wb = load_workbook(file_obj, data_only=True, read_only=True)
    ws = wb.active

    # Leer hasta fila 10 buscando la fila de cabecera
    header_row_idx = None
    idx_cod = idx_eps = idx_afp = None

    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    for i, row in enumerate(all_rows[:10]):
        heads = [str(c).strip().lower() if c else "" for c in row]
        if any("codigo" in h or "eps" in h for h in heads):
            header_row_idx = i
            for j, h in enumerate(heads):
                if "codigo" in h or "cód" in h:
                    idx_cod = j
                elif "eps" in h or "salud" in h:
                    idx_eps = j
                elif "afp" in h or "pension" in h or "pensión" in h:
                    idx_afp = j
            break

    # Fallback a posiciones fijas del archivo real (B=1, F=5, G=6)
    if header_row_idx is None: header_row_idx = 2  # fila 3, índice 2
    if idx_cod is None: idx_cod = 1
    if idx_eps is None: idx_eps = 5
    if idx_afp is None: idx_afp = 6

    maestro = {}
    for row in all_rows[header_row_idx + 1:]:
        if len(row) <= max(idx_cod, idx_eps, idx_afp):
            continue
        raw_cod, raw_eps, raw_afp = row[idx_cod], row[idx_eps], row[idx_afp]
        if not raw_cod:
            continue
        s = str(raw_cod).strip().replace(" ", "")
        if not s.isdigit():
            continue
        codigo = s.zfill(6)
        eps    = str(raw_eps).strip().upper() if raw_eps else ""
        afp    = str(raw_afp).strip().upper() if raw_afp else ""
        if codigo and eps and afp:
            maestro[codigo] = {"eps": eps, "afp": afp}

    wb.close()
    return maestro


def load_maestro() -> dict:
    """Devuelve el maestro de empleados con caché de 5 min.

    Orden de prioridad:
      1. OneDrive (si MAESTRO_URL está configurado)
      2. maestro.xlsx local
      3. Diccionario estático
    """
    global _maestro_cache, _maestro_cache_time
    if _maestro_cache is not None and (time.time() - _maestro_cache_time) < MAESTRO_CACHE_TTL:
        return _maestro_cache

    import requests as _req

    maestro = None

    # 1) OneDrive
    if MAESTRO_URL:
        try:
            url  = _onedrive_to_download(MAESTRO_URL)
            resp = _req.get(url, timeout=15)
            resp.raise_for_status()
            maestro = _parse_maestro_xlsx(io.BytesIO(resp.content))
            print(f"Maestro cargado desde OneDrive: {len(maestro)} empleados")
        except Exception as exc:
            print(f"Advertencia: no se pudo cargar maestro desde OneDrive: {exc}")

    # 2) Archivo local
    if maestro is None:
        local = os.path.join(BASE, "maestro.xlsx")
        if os.path.exists(local):
            try:
                with open(local, "rb") as fh:
                    maestro = _parse_maestro_xlsx(fh)
                print(f"Maestro cargado desde archivo local: {len(maestro)} empleados")
            except Exception as exc:
                print(f"Advertencia: no se pudo leer maestro.xlsx local: {exc}")

    # 3) Dict estático
    if maestro is None:
        maestro = MAESTRO_ESTATICO
        print(f"Maestro: usando diccionario estatico ({len(maestro)} empleados)")

    _maestro_cache      = maestro
    _maestro_cache_time = time.time()
    return maestro


MAESTRO_ESTATICO = {
    "040267": {"eps": "ALIANSALUD EPS",                                     "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "052231": {"eps": "E.P.S SANITAS",                                      "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "054755": {"eps": "E.P.S SANITAS",                                      "afp": "OLD MUTUAL FONDO DE PENSIONES OBLIGATORIAS"},
    "055512": {"eps": "E.P.S SANITAS",                                      "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "056374": {"eps": "FAMISANAR",                                          "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "056918": {"eps": "SALUD TOTAL S.A.",                                   "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "058834": {"eps": "NUEVA EPS",                                          "afp": "PORVENIR"},
    "059288": {"eps": "E.P.S SANITAS",                                      "afp": "PROTECCION"},
    "062858": {"eps": "E.P.S SANITAS",                                      "afp": "PROTECCION"},
    "073226": {"eps": "COMPENSAR ENTIDAD PROMOTORA DE SALUD",               "afp": "PORVENIR"},
    "073931": {"eps": "EPS SURA",                                           "afp": "PROTECCION"},
    "074173": {"eps": "EPS SURA",                                           "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "074490": {"eps": "EPS SURA",                                           "afp": "OLD MUTUAL FONDO DE PENSIONES OBLIGATORIAS"},
    "075064": {"eps": "EPS SURA",                                           "afp": "COLFONDOS"},
    "075115": {"eps": "FAMISANAR",                                          "afp": "COLFONDOS"},
    "075157": {"eps": "EPS SURA",                                           "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "075828": {"eps": "SALUD TOTAL S.A.",                                   "afp": "PROTECCION"},
    "076575": {"eps": "MUTUAL SER",                                         "afp": "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES"},
    "076845": {"eps": "E.P.S SANITAS",                                      "afp": "PROTECCION"},
    "077201": {"eps": "ALIANSALUD EPS",                                     "afp": "PORVENIR"},
    "077659": {"eps": "FAMISANAR",                                          "afp": "PROTECCION"},
    "077911": {"eps": "E.P.S SANITAS",                                      "afp": "PORVENIR"},
    "078082": {"eps": "EPS SURA",                                           "afp": "PORVENIR"},
    "078111": {"eps": "EPS SURA",                                           "afp": "PORVENIR"},
    "081214": {"eps": "COMPENSAR ENTIDAD PROMOTORA DE SALUD",               "afp": "PORVENIR"},
}

# EPS name → columna (índice 1-based) en formato2
EPS_COL = {
    "ALIANSALUD EPS":                         29,
    "EPS SURA":                               30,
    "SALUD TOTAL S.A.":                       31,
    "SALUD TOTAL SA":                         31,
    "SALUD TOTAL":                            31,
    "E.P.S SANITAS":                          32,
    "E.P.S. SANITAS":                         32,
    "SANITAS":                                32,
    "COMPENSAR ENTIDAD PROMOTORA DE SALUD":   33,
    "COMPENSAR":                              33,
    "FAMISANAR":                              34,
    "NUEVA EPS":                              35,
    "MUTUAL SER":                             36,
    "MUTUALSER":                              36,
}

AFP_PENSION_COL = {
    "COLFONDOS":                                              37,
    "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES":    38,
    "COLPENSIONES":                                           38,
    "OLD MUTUAL FONDO DE PENSIONES OBLIGATORIAS":             39,
    "OLD MUTUAL":                                             39,
    "SKANDIA":                                                39,
    "PORVENIR":                                               40,
    "PROTECCION":                                             41,
    "PROTECCIÓN":                                             41,
}

AFP_FSP_COL = {
    "COLFONDOS":                                              42,
    "ADMINISTRADORA COLOMBIANA DE PENSIONES COLPENSIONES":    43,
    "COLPENSIONES":                                           43,
    "OLD MUTUAL FONDO DE PENSIONES OBLIGATORIAS":             44,
    "OLD MUTUAL":                                             44,
    "SKANDIA":                                                44,
    "PORVENIR":                                               45,
    "PROTECCION":                                             46,
    "PROTECCIÓN":                                             46,
}

def normalize_code(v):
    return str(v).strip().replace(" ", "").zfill(6) if v else ""

def get_val(ws, col, row):
    return ws.cell(row=row, column=col).value

def set_val(ws, col, row, value):
    ws.cell(row=row, column=col).value = value

# Formato1 col → Formato2 col  (columnas directas, sin SS 41/42/43)
F1_TO_F2 = {
    2:2,  3:3,  6:4,  7:5,          # código, nombre, apellido 1, apellido 2
    8:6,  13:7,                       # 000050 días hábiles vac, 000051 días no hábiles vac
    17:8, 19:9,                       # 001050 Salario Integral, 001050 Salario..
    22:10, 23:11, 24:12, 25:13, 26:14,# 001060 recargo noc, HE diurnas, HE noc, HE dom/fest, recargo noc dom
    27:15, 28:16,                     # 001150 Inc enfermedad, 001151 Inc asumida
    29:17, 30:18,                     # 001170 Lic remunerada, 001300 Aux transporte
    31:19, 32:20, 33:21, 34:22, 35:23,# 100016 aux seg médico, 100019 aliment, 100020 transp, 100025 póliza, 100030 km
    39:24,                            # 100004 Prima Extralegal (posición distinta en F2)
    36:25, 37:26, 38:27,              # 001177 Aux inc asumida, 001178 Factor prest, 100052 Prima serv anticipada
    40:28,                            # DEVENGO total
    # cols 41, 42, 43 → ruteadas por entidad (EPS/AFP)
    44:47, 45:48, 46:49, 47:50, 48:51,# 003300 ret fuente, 200102 pen vol, 200200 afc bancolombia, 200203 colpatria, 200205 bco bogotá
    49:52, 50:53, 51:54, 52:55, 53:56, # 200302 seg vida, 200304 seg médico, 200305 fondexxom cart, 200306 serv, 200307 aportes
    54:57, 55:58, 56:59, 57:60, 58:61, 59:62, # 200311 seg med emp, 200313 ahorro, 200314 póliza fun, 200321 lib davivienda, 200322 compensar, 200324 loan
    60:63,                            # DEDUCCION total
    61:64, 62:65,                     # 999901 Neto a Pagar ×2
}


# ─── Routes ───────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(BASE, "index.html")

@app.route("/logosyp.png")
def logo():
    return send_from_directory(BASE, "logosyp.png")


@app.route("/process", methods=["POST"])
def process():
    if "archivo" not in request.files:
        return jsonify({"error": "No se recibió el archivo"}), 400

    f1_file = request.files["archivo"]
    if not f1_file.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "El archivo debe ser .xlsx"}), 400

    try:
        # 1. Leer formato1
        wb1 = load_workbook(f1_file, data_only=True)
        ws1 = wb1.active

        # 2. Detectar fila de inicio de datos (col 2 = código numérico)
        data_start = None
        for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
            v = row[1].value  # col B (índice 0 → col 2)
            if v and str(v).strip().replace(" ", "").isdigit() and len(str(v).strip()) >= 4:
                data_start = row[1].row
                break
        if data_start is None:
            return jsonify({"error": "No se encontraron filas de empleados en el reporte."}), 400

        # Recoger filas de empleados
        emp_rows = []
        for r in range(data_start, ws1.max_row + 1):
            v = get_val(ws1, 2, r)
            if v is None:
                break
            s = str(v).strip()
            if not s or not s[0].isdigit():
                break
            emp_rows.append(r)

        # 3. Abrir plantilla formato2 (preserva TODO el formato con openpyxl)
        template_path = os.path.join(BASE, "formato2.xlsx")
        wb2 = load_workbook(template_path)
        ws2 = wb2.active

        summary = []
        warnings = []
        f2_row = 25  # fila de inicio de datos en formato2

        # 4. Procesar cada empleado
        for f1r in emp_rows:
            raw_code = get_val(ws1, 2, f1r)
            code = normalize_code(raw_code)
            master = load_maestro().get(code)

            eps = master["eps"].upper() if master else None
            afp = master["afp"].upper() if master else None

            if not master:
                warnings.append(f"Sin maestro: {code}")

            # --- Copiar todas las columnas mapeadas ---
            for f1c, f2c in F1_TO_F2.items():
                val = get_val(ws1, f1c, f1r)
                if val is not None:
                    set_val(ws2, f2c, f2_row, val)
            # Código normalizado (sobrescribe col 2 con el formato correcto)
            set_val(ws2, 2, f2_row, code)

            # --- Valores SS ---
            v_salud   = get_val(ws1, 41, f1r) or 0
            v_pension = get_val(ws1, 42, f1r) or 0
            v_fsp     = get_val(ws1, 43, f1r) or 0

            # Rutear Salud
            if eps and v_salud:
                sc = EPS_COL.get(eps)
                if sc:
                    set_val(ws2, sc, f2_row, v_salud)
                else:
                    warnings.append(f"EPS no mapeada: '{eps}' ({code})")

            # Rutear Pensión
            if afp and v_pension:
                pc = AFP_PENSION_COL.get(afp)
                if pc:
                    set_val(ws2, pc, f2_row, v_pension)
                else:
                    warnings.append(f"AFP no mapeada: '{afp}' ({code})")

            # Rutear FSP
            if afp and v_fsp:
                fc = AFP_FSP_COL.get(afp)
                if fc:
                    set_val(ws2, fc, f2_row, v_fsp)
                else:
                    warnings.append(f"AFP/FSP no mapeada: '{afp}' ({code})")

            nombre = f"{get_val(ws1,6,f1r) or ''} {get_val(ws1,7,f1r) or ''}, {get_val(ws1,3,f1r) or ''}".strip()
            summary.append({
                "code": code,
                "nombre": nombre,
                "eps": eps or "—",
                "afp": afp or "—",
                "salud": v_salud,
                "pension": v_pension,
                "fsp": v_fsp,
                "ok": bool(master),
            })
            f2_row += 1

        # 5. Escribir a buffer y devolver
        buf = io.BytesIO()
        wb2.save(buf)
        buf.seek(0)

        filename = f"formato_ss_{date.today().isoformat()}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ), 200, {
            "X-Summary":  str(len(emp_rows)),
            "X-Warnings": str(len(warnings)),
            "X-Details":  "; ".join(warnings[:5]),
            "Access-Control-Expose-Headers": "X-Summary, X-Warnings, X-Details",
        }

    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


if __name__ == "__main__":
    print("=" * 55)
    print("  Servidor SS iniciado en http://localhost:5000")
    print("  Presiona Ctrl+C para detener")
    print("=" * 55)
    app.run(host="localhost", port=5000, debug=False)
